import time
from selenium.webdriver.common.by import By
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
import os
import openpyxl
from googletrans import Translator
import requests
import io
import tempfile
from requests_html import HTMLSession

# Create a session
session = HTMLSession()

# Set up the Selenium webdriver (replace the path with the location of your chromedriver)
ser = Service('/usr/local/bin/chromedriver')

opt=Options()
opt.add_experimental_option("debuggerAddress","localhost:9222")
driver = webdriver.Chrome(service=ser,options=opt)
# Navigate to the website

#invalid
url='https://m.tb.cn/h.gSMweC3?tk=5Izv3biQRya'

#valid
url='https://www.goofish.com/item?spm=a21ybx.item.itemCnxh.6.41bc3da6odCxYV&id=820650751634&categoryId=0'


def translate_text(text, src='zh-CN', dest='en'):
    """
    Translates the given text from the source language to the destination language.

    Args:
    text (str): The text to be translated.
    src (str): The source language code (default is 'zh-CN' for Chinese).
    dest (str): The destination language code (default is 'en' for English).

    Returns:
    str: The translated text.
    """
    translator = Translator()
    translation = translator.translate(text, src=src, dest=dest)
    return translation.text


def get_url(data_field):
    if isinstance(data_field, str):
        url_start = data_field.find("https://")
        if url_start != -1:
            url_end = data_field.find(" ", url_start)
            if url_end != -1:
                url = data_field[url_start:url_end]
                return url
            else:
                return data_field
    return None

def page_is_loading(driver):
    while True:
        x = driver.execute_script("return document.readyState")
        if x == "complete":
            return True
        else:
            yield False

def get_info(url):

    if (url !=None):
        global driver
        print('url= ',url)
        driver.get(url)

    return info

def check_url(url):
    is_valid_url = False

    if (url !=None):
        global driver
        print('url= ',url)
        driver.get(url)

        is_valid_url=False

        while not page_is_loading(driver):
            continue
        print(' page all loaded ')

        # while True:
        #     try:
        #         # Check if the page has finished loading
        #         is_page_loaded = driver.execute_script("return document.readyState == 'complete'")
        #         if is_page_loaded:
        #             break
        #         else:
        #             time.sleep(0.5)  # Wait for 0.5 seconds before checking again
        #     except:
        #         # If any exception occurs, the page is considered not loaded
        #         is_page_loaded = False
        #         break
        time.sleep(3)
        current_url = driver.current_url
        print(current_url)

        error_url = "https://m.tb.cn/scanError.htm"
        if error_url in current_url:
            print("Error: The system is forwarding to the error page.")
            return False


        page_content = driver.page_source
        description=''
        price_value=''
        title=''

        # Check if the page content includes the text "卖掉了"
        if "卖掉了" in page_content:
            return "卖掉了"
        elif "已下架" in page_content:
            return "已下架"
        else:
            # div_element = driver.find_element(By.CSS_SELECTOR, "price--Ls68DZ8a")
            # print("price",div_element.text)


            try:
                price_element = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div[1]/div[2]/div[2]/div[2]/div[1]/div/div[2]")
                price_value = price_element.get_attribute("textContent")
                print("price", price_value)
                price_element = driver.find_element(By.XPATH,"/ html / body / div[1] / div[2] / div[1] / div[2] / div[2] / div[4] / div / span / span / span"
                )
                desc = price_element.get_attribute("textContent")
                title = translate_text(desc)

                print("desc", desc)

                print("eng", title)

            except:
                print("desc not found")


            for i in range(2,6):
                path_link = "/ html / body / div[1] / div[2] / div[1] / div[2] / div[2] / div[4] / div / span / span[" + str(i) + "] / span"
                print(path_link)

                try:

                    price_element = driver.find_element(By.XPATH,path_link  )
                    desc2 = price_element.get_attribute("textContent")
                    print("desc2", desc2)
                    description = description + "\n" + translate_text(desc2)
                    print("eng", description)
                except:
                    print("desc2 not found")

            line=["正常",price_value,title,description]

            return line

    return is_valid_url


def get_photo(row_num):
    global driver

    # Wait for the page to load
    # wait = WebDriverWait(driver, 10)
    # wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))

    div_element = driver.find_element(By.CSS_SELECTOR, "div.ant-carousel.css-1uq9j6g")
    print(div_element)

    # Find all the image elements within the div
    img_elements = div_element.find_elements(By.TAG_NAME, "img")

    # Download and save the images
    i=0
    print(len(img_elements))
    max=(len(img_elements)/2) + 1

    for img_element in img_elements:
        i+=1
        img_src = img_element.get_attribute("src")

        if (i<= max ) & (i>1):
            print(img_src)
            response = session.get(img_src)

            #response = requests.get(img_src)
            content_type = response.headers.get('Content-Type')

            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
                temp_file.write(response.content)
                img = Image.open(temp_file.name)
                img_id=(str(row_num) + str(i) ).zfill(4)
                img.save(f"./280824-bag/IMG_{img_id}.jpg")

    return int(max)

def main():
    workbook = openpyxl.load_workbook('280824-bag.xlsx')

    # Select the 'Aug24-Earrings' sheet
    worksheet = workbook['Sheet1']

    # Iterate through the rows and update the values in the 'C' column
    for row in range(2, worksheet.max_row + 1):
        url = get_url(worksheet.cell(row=row, column=2).value)
        result = check_url(url)

        if isinstance(result, str):
            worksheet.cell(row=row, column=3).value = result
            print("The return value is a string:", result)
        elif isinstance(result, list):
            worksheet.cell(row=row, column=3).value = result[0]

            worksheet.cell(row=row, column=4).value = '220824-' + str(row) + ' ' + result[2]
            worksheet.cell(row=row, column=9).value = result[3]
            print(type(result[1]))

            if (result[1] !=""):
                worksheet.cell(row=row, column=6).value = int(int(result[1]) * 0.2)


            img_src= str(row) + '2;' + str(row) + str(get_photo(row))
            worksheet.cell(row=row, column=8).value = img_src


            print("The return value is a list:", result[0])
        else:
            print("The return value is of an unexpected type:", type(result))

        if (worksheet.cell(row=row, column=2).value is None):
            break

        workbook.save('240824-Earrings.xlsx')

    # Save the changes to the Excel file




main()

# if check_url(url):
#     get_photo()
#
# driver.quit()