import time
from selenium.webdriver.common.by import By
from PIL import Image
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
import os
import openpyxl
from googletrans import Translator
import requests
import io
import tempfile
from requests_html import HTMLSession

import mysql.connector
import shopify_api
import sys
import os

import socket
import time
import re
from bs4 import BeautifulSoup


def send_data(sock, data):
    try:
        sock.sendall(data)
    except BrokenPipeError:
        print("Connection closed. Reconnecting...")
        # Reconnect logic here
        sock = reconnect()  # Define reconnect function as needed
        sock.sendall(data)

# Get the directory of the current script
current_dir = os.path.dirname(os.path.abspath(__file__))

# Get the parent directory of the config directory
parent_dir = os.path.dirname(current_dir)

# Add the config directory to the system path
sys.path.append(parent_dir)
import config

def create_connection():
    return mysql.connector.connect(
        host='hkwaishing.com',
        user='waishing_trendy',
        password='Socool666',
        database='waishing_binance'
    )


connection = create_connection()
cursor = connection.cursor()

# mydb = mysql.connector.connect(
#     host=config.host,
#     user=config.user,
#     passwd=config.passwd,
#     database=config.database
# )
#
# cursor = mydb.cursor()

# Create a session
session = HTMLSession()

# Set up the Selenium webdriver (replace the path with the location of your chromedriver)
ser = Service('/usr/local/bin/chromedriver')

opt = Options()
opt.add_experimental_option("debuggerAddress", "localhost:9222")
driver = webdriver.Chrome(service=ser, options=opt)










def translate_text(text, src='zh-CN', dest='en'):
    """
    Translates the given text from the source language to the destination language.

    Args:
    text (str): The text to be translated.
    src (str): The source language code (default is 'zh-CN' for Chinese).
    dest (str): The destination language code (default is 'en' for English).

    Returns:
    str: The translated text.
    """
    translator = Translator()
    translation = translator.translate(text, src=src, dest=dest)
    return translation.text


def get_url(data_field):



    if isinstance(data_field, str):
        if (data_field in 'auctions'):
            return data_field
        url_start = data_field.find("https://")
        if url_start != -1:
            url_end = data_field.find(" ", url_start)
            if url_end != -1:
                url = data_field[url_start:url_end]
                return url
            else:
                return data_field
    return None


def get_japan_product(status):
    global driver

    title_element = driver.find_element(By.CLASS_NAME, 'ProductTitle__text')

    # Get the text value
    title_value = title_element.text

    title = translate_text(title_value, 'ja', 'en')


    price_element = driver.find_element(By.CLASS_NAME, 'Price__value')

    # Get the text value
    japanese_price = price_element.text
    print(japanese_price)
    # Extract the value using regex
    if "税込" in japanese_price:
        # Extract the value with tax
        match = re.search(r'税込 (\d{1,3}(?:,\d{3})*) 円', japanese_price)
    else:
        # Extract the value without tax
        match = re.search(r'(\d{1,3}(?:,\d{3})*)円', japanese_price)


    if match:
        jpy_value_str = match.group(1)
        jpy_value = int(jpy_value_str.replace(',', ''))  # Convert to integer
        print(f"Extracted JPY Value: {jpy_value}")

        # URL of a reliable exchange rate provider
        url = 'https://www.x-rates.com/calculator/?from=JPY&to=CNY'

        # Make the request
        response = requests.get(url)
        soup = BeautifulSoup(response.text, 'html.parser')

        # Find the exchange rate value
        rate_element = soup.find('span', class_='ccOutputRslt')
        if rate_element:
            rate_text = rate_element.text.strip()  # Get the text and remove any leading/trailing spaces
            # Extract just the numeric part
            try:
                rate_value = float(rate_text.split()[0])  # Convert to float
                print(f"Current exchange rate from JPY to CNY: {rate_value}")
            except ValueError:
                print("Failed to convert the exchange rate to a float.")
        else:
            print("Failed to retrieve exchange rate. The structure of the page may have changed.")



        # Convert JPY to RMB
        rmb_value = float(jpy_value * rate_value)
        print(f"Value in CNY: ${rmb_value:.2f}")
    else:
        print("No match found.")


    # Print the extracted price value

    line = [status,rmb_value,title,title]
    return line

def page_is_loading(driver):
    while True:
        x = driver.execute_script("return document.readyState")
        if x == "complete":
            return True
        else:
            yield False

def get_info(url):

    if (url !=None):
        global driver
        print('url= ',url)
        driver.get(url)

    return info

def check_url(url,mode):
    is_valid_url = False

    if (url !=None):
        global driver
        print('url= ',url)
        driver.get(url)

        is_valid_url=False

        while not page_is_loading(driver):
            continue
        #print(' page all loaded ')

        # while True:
        #     try:
        #         # Check if the page has finished loading
        #         is_page_loaded = driver.execute_script("return document.readyState == 'complete'")
        #         if is_page_loaded:
        #             break
        #         else:
        #             time.sleep(0.5)  # Wait for 0.5 seconds before checking again
        #     except:
        #         # If any exception occurs, the page is considered not loaded
        #         is_page_loaded = False
        #         break
        time.sleep(3)
        current_url = driver.current_url
        print(current_url)

        error_url = "https://m.tb.cn/scanError.htm"
        if error_url in current_url:
            print("Error: The system is forwarding to the error page.")
            return False


        page_content = driver.page_source
        description=''
        price_value=''
        title=''
        status=''

        # Check if the page content includes the text "卖掉了"
        if "卖掉了" in page_content:
            status= "卖掉了"
        elif "已下架" in page_content:
            status= "已下架"
        elif "宝贝被删掉了" in page_content:
            status= "宝贝被删掉了"
        elif "糟糕！宝贝被删掉了" in page_content:
            status="糟糕！宝贝被删掉了"
        elif "このオークションは終了しています" in page_content:
            status= get_japan_product("終了")
        elif "今すぐ落札" in page_content:
            status= get_japan_product("即決")
        elif ">入札<" in page_content:
            status= get_japan_product("入札")
        elif "我想要" in page_content:
            status= "正常"
        else:
            return "不明"




        try:

            price_element = driver.find_element(By.CLASS_NAME, 'price--Ls68DZ8a ')
            price_value = price_element.text
            print('BY Class ',price_value)

            # price_element = driver.find_element(By.XPATH,"/html/body/div[1]/div[2]/div[1]/div[2]/div[2]/div[2]/div[1]/div/div[2]")
            # price_value = price_element.get_attribute("textContent")
            # print("price", price_value)
        except Exception as error:
            print(error)


        try:
            desc_element = driver.find_element(By.XPATH,"/ html / body / div[1] / div[2] / div[1] / div[2] / div[2] / div[4] / div / span / span / span"
            )

            desc = desc_element.get_attribute("textContent")
            title = translate_text(desc)

            print("desc", desc)

            print("eng", title)

        except Exception as error:
            print(error)

        try:
            span_elements = driver.find_elements(By.CSS_SELECTOR, 'span.desc--WLgQcGKD span')

            title_value = span_elements[0].text
            title = translate_text(title_value)
            unique_values = set(span.text for span in span_elements)

            # Combine the unique values into a single string
            desc = " + ".join(unique_values)

            # Print the combined result
            print(description)  # Output: A + B (order may vary)

            description = translate_text(desc)


        except Exception as error:
            print(error)





        except Exception as error:

            print(error)




        line=[status,price_value,title,description,current_url]

        return line



    return is_valid_url


def create_directory(directory_name):
    """Create a directory if it does not exist."""
    if not os.path.exists(directory_name):
        os.makedirs(directory_name)
        print(f"Directory '{directory_name}' created.")
    else:
        print(f"Directory '{directory_name}' already exists.")


def get_photo(row_num,sheet_name):
    global driver
    max=0
    i=0
    all_img=''
    # Wait for the page to load
    # wait = WebDriverWait(driver, 10)
    # wait.until(EC.presence_of_element_located((By.TAG_NAME, 'body')))
    try:

        create_directory(sheet_name)

        img_elements = driver.find_elements(By.CSS_SELECTOR, '.ant-image-img.css-vryruh')

        # img_elements = WebDriverWait(driver, 10).until(
        #     EC.presence_of_all_elements_located((By.CSS_SELECTOR, '.ant-image-img.css-vryruh'))
        # )



        # Extract and print the src attributes of all image elements
        i = 0
        img_links = [img.get_attribute('src') for img in img_elements]


        for link in img_links:
            print(link)
            i += 1
            all_img += link + ','
            response = session.get(link)

            # response = requests.get(img_src)
            content_type = response.headers.get('Content-Type')

            with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
                temp_file.write(response.content)
                img = Image.open(temp_file.name)
                img_id = (str(row_num).zfill(2) + str(i).zfill(2))
                img.save(f"./{sheet_name}/IMG_{img_id}.jpg")

        max=i

        #div_element = driver.find_element(By.CSS_SELECTOR, "div.ant-carousel.css-1uq9j6g")

        # div_element = driver.find_element(By.CSS_SELECTOR, ".ant-image-img.css-vryruh")
        #
        # print('start get_photo taobao')
        #
        # print(div_element)
        #
        # # Find all the image elements within the div
        # img_elements = div_element.find_elements(By.TAG_NAME, "img")
        #
        # # Download and save the images
        # i = 0
        # print(len(img_elements))
        #
        #
        # max = (len(img_elements) / 2) + 1
        #
        # for img_element in img_elements:
        #     i += 1
        #     img_src = img_element.get_attribute("src")
        #
        #     if (i <= max) & (i > 1):
        #         print(i, "-", max, img_src)
        #         all_img += img_src + ','
        #         response = session.get(img_src)
        #
        #         # response = requests.get(img_src)
        #         content_type = response.headers.get('Content-Type')
        #
        #         with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
        #             temp_file.write(response.content)
        #             img = Image.open(temp_file.name)
        #             img_id=(str(row_num).zfill(2) + str(i).zfill(2) )
        #             img.save(f"./{sheet_name}/IMG_{img_id}.jpg")


    except Exception as error:
        print('Taobao get_photo Error - ' + error)

        try:
            # Find all elements with the class 'ProductImage__inner'
            print('start get_photo Japan')
            div_element = driver.find_elements(By.CLASS_NAME, 'ProductImage__inner')
            i = 0

            # Extract image URLs
            unique_image_urls = []
            seen_urls = set()

            # Create a directory named after the translated title (sanitize it for folder names)
            directory_name = sheet_name.replace('/', '_').replace('\\', '_').replace(':', '_')
            os.makedirs(directory_name, exist_ok=True)

            for container in div_element:
                img = container.find_element(By.TAG_NAME, 'img')
                if img:
                    src = img.get_attribute('src')
                    if src not in seen_urls:
                        seen_urls.add(src)
                        unique_image_urls.append(src)

            # Print the extracted image URLs
            for img_url in unique_image_urls:
                all_img += img_url + ','
                response = session.get(img_url)
                i += 1

                # response = requests.get(img_src)
                content_type = response.headers.get('Content-Type')

                with tempfile.NamedTemporaryFile(delete=False, suffix=".jpg") as temp_file:
                    temp_file.write(response.content)
                    img = Image.open(temp_file.name)
                    img_id=(str(row_num).zfill(2) + str(i).zfill(2) )
                    img.save(f"./{sheet_name}/IMG_{img_id}.jpg")

            max=i
            print('Japan', all_img, max)

        except Exception as error:
            print(error)




    return int(max), all_img



def update_product_status(sheet_name):

    global cursor, mydb

    try:
        connection = create_connection()
        cursor = connection.cursor()
    except Exception as error:
        print(error)

    # Query to fetch data with pagination
    #cursor.execute(f"SELECT URL, STATUS, PRODUCT_ID FROM PRODUCT WHERE STATUS='不明' ")
    # Define the SQL query as a string
    sql = """
        SELECT URL, STATUS, PRODUCT_ID 
        FROM PRODUCT 
        WHERE SHEET_NAME = %s 
        AND (NOW()  > CHECK_DATE + INTERVAL 1 DAY OR CHECK_DATE = '0000-00-00 00:00:00')
    """

    # Execute the query
    print(sql)
    cursor.execute(sql, (sheet_name,))

    results = cursor.fetchall()
    print(results)

    for row in results:
        url, status, product_id = row
        print(row)

        try:
            connection = create_connection()
            cursor = connection.cursor()
        except Exception as error:
            print(error)

        result = check_url(url,'no')

        if isinstance(result, str):
            new_status=result
        elif isinstance(result, list):
            new_status = result[0]
        else:
            new_status='不明'

        if new_status == status:
            print('Not Change ', new_status)
            query = """
                        UPDATE PRODUCT 
                        SET CHECK_DATE = CURRENT_TIMESTAMP 
                        WHERE PRODUCT_ID = %s
                        """

            print(query)

            try:
                cursor.execute(query, ( product_id,))
                connection.commit()
            except mysql.connector.OperationalError as e:
                print(f"OperationalError: {e}")
                cursor.close()
                connection.close()
                time.sleep(2)  # Wait before reconnecting
                connection = create_connection()
                cursor = connection.cursor()
                continue  # Retry the current iteration

        elif new_status != '不明':
            print('Status Changed to ',new_status)
            current_timestamp = "CURRENT_TIMESTAMP"  # MySQL function for current timestamp

            # Update query
            query = """
                        UPDATE PRODUCT 
                        SET STATUS = %s, UPDATED_DATE = CURRENT_TIMESTAMP , CHECK_DATE = CURRENT_TIMESTAMP 
                        WHERE PRODUCT_ID = %s
                        """

            print(query)

            try:
                cursor.execute(query, (new_status, product_id))
                connection.commit()
                shopify_api.search_products_by_tag_graphql(product_id)

            except mysql.connector.OperationalError as e:
                print(f"OperationalError: {e}")
                cursor.close()
                connection.close()
                time.sleep(2)  # Wait before reconnecting
                connection = create_connection()
                cursor = connection.cursor()
                continue  # Retry the current iteration

            except Exception as e:
                print(f"Error: {e}")

        else:
            print('不明 status')





def insert_product_db(date_code):

    global cursor

    excel_file = date_code + '.xlsx'

    workbook = openpyxl.load_workbook(excel_file)

    # Select the 'Aug24-Earrings' sheet
    worksheet = workbook['Sheet1']

    # Iterate through the rows and update the values in the 'C' column
    for row in range(1, worksheet.max_row + 1):
        PRODUCT_ID= date_code + '-' + str(row)
        sql="SELECT URL, STATUS, PRODUCT_ID FROM PRODUCT WHERE PRODUCT_ID = %s "
        cursor.execute(sql, (PRODUCT_ID,))  # Note the comma to create a tuple
        results = cursor.fetchall()


        try:
            connection = create_connection()
            cursor = connection.cursor()
        except Exception as error:
            print(error)

        if results:
            print(PRODUCT_ID, " product id exist")
        else:

            url = get_url(worksheet.cell(row=row, column=2).value)
            result = check_url(url,"FULL")

            if isinstance(result, str) and result !='不明':

                print("The return value is a string:", result)
                # sql = 'INSERT INTO PRODUCT (PRODUCT_ID,STATUS,URL) VALUES ("' + date_code + str(row) + '","' + str(result) + '","' + url + '")'
                # print(sql)
                max_num, all_img= get_photo(row,date_code)

                query = """
                            INSERT INTO PRODUCT (PRODUCT_ID,STATUS,URL,PHOTO, SHEET_NAME)
                            VALUES ( %s, %s, %s, %s, %s)
                            """
                print(query)

                try:
                    cursor.execute(query, (PRODUCT_ID, str(result), url, str(all_img),date_code))
                    connection.commit()
                except Exception as error:
                    print(error)


            elif isinstance(result, list) and result[0] !='不明':
                print('If list ----',result)
                #line=["正常",price_value,title,description,current_url]
                price=0
                cost=0
                status = result[0]


                if (len(result) > 1 ):
                    if (result[1] != ""):
                        price = int(float(result[1]) * 0.2)
                        cost= int(float(result[1]) * 0.14)

                if (len(result) > 2 ):
                    title = PRODUCT_ID + ' ' + str(result[2])


                if (len(result)>3):
                    desc = result[3]

                if (len(result)>4):
                    url = result[4]



                print(type(result[1]))



                max_num, all_img= get_photo(row,date_code)

                img_src = str(row).zfill(2) + '01;' + str(row).zfill(2) + str(max_num).zfill(2)
                print(all_img)

                escaped_desc = desc.replace('"', '\\"')  # Escape double quotes
                escaped_title = title.replace('"', '\\"')  # Escape double quotes

                # Construct the SQL statement
                sql = (
                        'INSERT INTO PRODUCT (PRODUCT_ID, PHOTO, STATUS, URL, IMAGE_SRC, TITLE, PRICE, BODY, COST, SHEET_NAME) '
                        'VALUES ("' + PRODUCT_ID + '", "' + str(all_img) + '", "' + str(status) + '", "' +
                        str(url) + '", "' + str(img_src) + '", "' + escaped_title + '", ' + str(price) + ', "' +
                        PRODUCT_ID + ' P ' + str(cost) + ' ' + escaped_desc + '", ' + str(cost) + ', "' + str(
                    date_code) + '")'
                )

                #sql = 'INSERT INTO PRODUCT (PRODUCT_ID,PHOTO,STATUS,URL,IMAGE_SRC,TITLE,PRICE,BODY,COST,SHEET_NAME) VALUES ("' + PRODUCT_ID + '","' + str(all_img) +  '","' + str(status) + '","' + str(url) + '","' + str(img_src) + '","' + str(title) + '",' + str(price) + ',"' + PRODUCT_ID + '  P' + str(cost) + ' ' + str(desc) + '",' + str(cost) +   ',"' + str(date_code) + '" )'
                print('here' + sql)

                try:
                    cursor.execute(sql)
                    connection.commit()
                except Exception as error:
                    print(error)

                print("The return value is a list:", status)

            else:
                print("The return value is of an unexpected type:", type(result))






        if (worksheet.cell(row=row, column=2).value is None):
            break




def save_excel():

    date_code = 'sep24-earrings'
    excel_file = 'sep24-earrings.xlsx'
    workbook = openpyxl.load_workbook(excel_file)

    # Select the 'Aug24-Earrings' sheet
    worksheet = workbook['Sheet1']

    # Iterate through the rows and update the values in the 'C' column
    for row in range(1, worksheet.max_row + 1):

        try:
            url = get_url(worksheet.cell(row=row, column=2).value)
            result = check_url(url,"FULL")

            if isinstance(result, str):
                worksheet.cell(row=row, column=3).value = result
                print("The return value is a string:", result)
            elif isinstance(result, list):
                worksheet.cell(row=row, column=3).value = result[0]

                worksheet.cell(row=row, column=4).value = date_code + str(row) + ' ' + result[2]
                worksheet.cell(row=row, column=9).value = result[3]
                print(type(result[1]))

                if (result[1] !=""):
                    worksheet.cell(row=row, column=6).value = int(int(result[1]) * 0.14)

                max_num, all_img= get_photo(row,2,date_code)

                img_src = str(row).zfill(2) + '01;' + str(row).zfill(2) + str(max_num).zfill(2)

                #img_src= str(row) + '2;' + str(row) + str(get_photo(row))
                worksheet.cell(row=row, column=8).value = img_src


                print("The return value is a list:", result[0])
            else:
                print("The return value is of an unexpected type:", type(result))

        except Exception as error:
            print(error)


        if (worksheet.cell(row=row, column=2).value is None):
            break

        workbook.save(excel_file)

    # Save the changes to the Excel file
if __name__ == "__main__":

    print('start ')
#main()
#check_url('https://m.tb.cn/h.govfdb3?tk=zKhO3TVcqRD','no')
    #insert_product_db('191024')

    #update_product_status('191024')
# url='https://page.auctions.yahoo.co.jp/jp/auction/1153180042'
# result = check_url(url,"no")
# print(result)
# max_num, all_img = get_photo(6, 2,'1809-japan')

#

# if check_url(url):
#     get_photo()
#
# driver.quit()